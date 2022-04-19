import openpyxl
from openpyxl.styles import Font
from datetime import datetime

nowTime = datetime.now().strftime('%d''-%m''-%Y''-%H:%M:%S')


def createExcel(appName, vulnPacket, vulnName, descURI):
    wb = openpyxl.Workbook()
    sheet = wb.active

    a1 = sheet.cell(row=1, column=1)
    a1.value = "APP NAME"
    a1.font = Font(bold=True)

    b1 = sheet.cell(row=1, column=2)
    b1.value = "VULNERABILITY PACKET"
    b1.font = Font(bold=True)

    c1 = sheet.cell(row=1, column=3)
    c1.value = "VULNERABILITY NAME"
    c1.font = Font(bold=True)

    d1 = sheet.cell(row=1, column=4)
    d1.value = "DESCRIPTION URI"
    d1.font = Font(bold=True)

    rowNumber = 2
    for i in range(len(vulnPacket)):
        columnAppName = sheet.cell(row=rowNumber, column=1)
        columnAppName.value = appName

        columnPacket = sheet.cell(row=rowNumber, column=2)
        columnPacket.value = vulnPacket[i]

        columnName = sheet.cell(row=rowNumber, column=3)
        columnName.value = vulnName[i]

        columnDesc = sheet.cell(row=rowNumber, column=4)
        columnDesc.value = descURI[i]

        rowNumber += 1

    wb.save(f"reports/{appName.replace('/', '_')}-{nowTime}-snyk-reports.xlsx")
